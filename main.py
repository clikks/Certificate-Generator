#!/usr/bin/env python3
#-*- coding: utf-8 -*-

import os
import sys

from plumbum import cli
from plumbum.cmd import openssl
from openpyxl import Workbook, load_workbook

from devicelist import get_citylist, certificate_dir, xlsx


class OptParser(cli.Application):
    PROGNAME = 'Certificates Generator'
    VERSION = 'v0.1'

    @cli.switch(["a", "-amount"], argtype=int, mandatory=True, help='The total number of certificates that need to be generated')
    def cert_num(self, amount):
        self._amount = amount

    @cli.switch(["c", "-city"], argtype=str, requires = ["a","-amount"], help='Specify the city that needs to generate certificates')
    def cert_city(self, city):
        self._city = (city, get_citylist(city))

    def main(self):
        inwb = load_workbook(xlsx)
        try:
            del inwb[str(self._city[0])]
            sheet = inwb.create_sheet(0)
            sheet.title = '{}'.format(self._city[0])    
        except KeyError:
            sheet = inwb.create_sheet(0)
            sheet.title = '{}'.format(self._city[0])
        except AttributeError:
            sheet = inwb.create_sheet(0)
            sheet.title = 'amount{}'.format(self._amount)
        global key_dict
        key_dict = dict()
        try:
            cert_gen(self._amount, self._city)
        except AttributeError:
            cert_gen(self._amount, None)
        # print(key_dict)
        for index, sn in enumerate(key_dict):
            column = int(index)+1
            sheet['A{}'.format(column)].value = sn
            sheet['B{}'.format(column)].value = key_dict.get(sn)[0]
            sheet['C{}'.format(column)].value = key_dict.get(sn)[1]
        inwb.save(xlsx)
        print('已完成，请查看excel.')


def build_key(amount,snlist):
    for sn in snlist.keys():
        os.mkdir(sn.upper())
        print('{} creation...'.format(sn))
        path = os.path.join(os.getcwd(),sn)
        priv_key = os.path.join(path,"rsa_private_key.pem")
        pkcs8_key = os.path.join(path,"rsa_private_key_pkcs8.pem")
        pub_key = os.path.join(path, "public_key.pem")
        gen_priv = openssl["genrsa","-out",priv_key,"2048"]
        gen_priv_pkcs8 = openssl["pkcs8","-topk8","-inform","PEM","-in",priv_key,"-outform","PEM","-nocrypt","-out",pkcs8_key]
        gen_pub = openssl["rsa","-in",priv_key,"-pubout","-out",pub_key]
        gen_priv()
        gen_priv_pkcs8()
        gen_pub()
        mac = snlist.get(sn)
        get_pub_key(pub_key,sn.upper(),mac)
    if len(snlist) < amount:
        build_amount_key(amount,snlist)


def build_amount_key(amount,snlist):
    if len(snlist) == 0:
        if not os.path.exists('tmp'):
            os.mkdir('tmp')
        os.chdir('tmp')
    for num in range(1,amount-len(snlist)+1):
        os.mkdir(str(num))
        print('{} creation...'.format(num))
        path = os.path.join(os.getcwd(), str(num))
        priv_key = os.path.join(path,"rsa_private_key.pem")
        pkcs8_key = os.path.join(path,"rsa_private_key_pkcs8.pem")
        pub_key = os.path.join(path, "public_key.pem")
        gen_priv = openssl["genrsa","-out",priv_key,"2048"]
        gen_priv_pkcs8 = openssl["pkcs8","-topk8","-inform","PEM","-in",priv_key,"-outform","PEM","-nocrypt","-out",pkcs8_key]
        gen_pub = openssl["rsa","-in",priv_key,"-pubout","-out",pub_key]
        gen_priv()
        gen_priv_pkcs8()
        gen_pub()
        get_pub_key(pub_key,num,0)
        

def get_pub_key(pub_key_path,sn,mac):
    with open(pub_key_path) as f:
        data = f.readlines()
    key = ''
    for line in data[1:-1]:
        key = key + line.replace('\n','')
    print('公钥生成完毕：')
    print(key)
    key_dict[sn] = [mac,key]
   


def cert_gen(amount, city):
    if city is not None:
        cityname,snlist = city
        print(snlist)
        if snlist is None:
            print('指定城市数据不存在')
        else:
            if len(snlist) > amount:
                print('指定生成证书个数小于城市SN数据个数，将按照SN生成证书！')
            os.chdir(certificate_dir)
            if os.path.exists(cityname) and len(os.listdir(cityname)) == 0:
                os.chdir(cityname)
                print('目录存在并且为空，开始生成证书...')
                build_key(amount,snlist)
            elif os.path.exists(cityname) and len(os.listdir(cityname)) != 0:
                print('{}已有SN对应目录，请检查目录并清空后重新运行。'.format(cityname))
            else:
                print('城市目录不存在，创建{}目录'.format(cityname))
                os.mkdir(cityname.lower())
                os.chdir(cityname.lower())
                build_key(amount,snlist)
    else:
        os.chdir(certificate_dir)
        build_amount_key(amount,[])



if __name__ == '__main__':
    OptParser()
