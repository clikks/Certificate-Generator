#!/usr/bin/env python3
# encoding: utf-8
# __author__= 'xul'

import os
import logging
from datetime import datetime
from collections import OrderedDict


from plumbum.cmd import openssl
from openpyxl import load_workbook
from devicelist import certificate_dir, get_citylist, xlsx


logger = logging.getLogger(__name__)
logger.setLevel(level=logging.INFO)
console = logging.StreamHandler()
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
console.setFormatter(formatter)
logger.addHandler(console)


def close_log():
    logger.removeHandler(console)


def load_infile(cityname, infile_path):
    """
    读取存储存储sn和mac的excel文件中的数据进行处理
    """
    # device_data = dict()
    device_data = OrderedDict()
    inwb = load_workbook(infile_path)
    sheet_names = inwb.sheetnames
    if cityname in sheet_names:
        sheet = inwb[cityname]
        for sn, mac in sheet.rows:
            device_data[sn.value] = mac.value
        return device_data
    else:
        print('{}中{}数据不存在，请检查表格！'.format(os.path.split(infile_path)[1], cityname))
        exit(-1)


def get_device_data(cityname, amount, infile):
    if cityname:
        if infile:
            device_data = load_infile(cityname, infile)
            logger.info('Read device_data from {} successfully:\n{}'.format(os.path.split(infile)[1], device_data))
        else:
            device_data = OrderedDict(get_citylist(cityname))
            logger.info('Read device_data from devicelist successfully:\n{}'.format(device_data))
        if amount:
            cert_num = int(amount) - len(device_data)
            for i in range(1, cert_num+1):
                device_data[str(i)] = 'Null'
            logger.debug(device_data)
        return device_data
    elif amount:
        device_data = OrderedDict()
        for i in range(1, int(amount)+1):
            device_data[str(i)] = 'Null'
        # device_data = {str(i): 'Null' for i in range(1, int(amount)+1)}
        logger.debug(device_data)
        return device_data
    else:
        return None


def create_city_dir(cityname):
    if cityname:
        cityname_dir = os.path.join(certificate_dir, cityname)
        try:
            os.mkdir(cityname_dir)
            logger.info('创建 {} 目录成功!'.format(cityname))
        except FileExistsError:
            if len(os.listdir(cityname_dir)) != 0:
                logger.info('创建 {} 目录失败，目录已存在并且非空。'.format(cityname))
    else:
        dirname = datetime.strftime(datetime.now(), '%Y-%m-%d-%H-%M-%S')
        cityname_dir = os.path.join(certificate_dir, dirname)
        os.mkdir(cityname_dir)
        logger.info('Create {} directory success！'.format(dirname))
    os.chdir(cityname_dir)
    logger.info('Change to {} directory success！'.format(cityname_dir))    
    return cityname_dir


def key_gen(cityname_dir, device_data):
    logger.info('current directory is {}'.format(os.getcwd()))
    key_dict = OrderedDict()
    for sn, mac in device_data.items():
        os.mkdir(sn)
        device_dir = os.path.join(cityname_dir, sn)
        priv_key_path = os.path.join(device_dir, "rsa_private_key.pem")
        priv_pkcs8_key_path = os.path.join(device_dir, "rsa_private_key_pkcs8.pem")
        pub_key_path = os.path.join(device_dir, "public_key.pem")

        gen_priv_key = openssl["genrsa", "-out", priv_key_path, "2048"]
        gen_priv_pkcs8_key = openssl["pkcs8", "-topk8", "-inform", "PEM", "-in", priv_key_path, "-outform", "PEM", "-nocrypt", "-out", priv_pkcs8_key_path]
        gen_pub_key = openssl["rsa", "-in", priv_key_path, "-pubout", "-out", pub_key_path]

        gen_priv_key()
        logger.info('Successfully generated the {} private key'.format(sn))
        gen_priv_pkcs8_key()
        logger.info('successfully generate the {} pksc8 format private key.'.format(sn))
        gen_pub_key()
        logger.info('successfully generate the {} public key.'.format(sn))

        with open(pub_key_path) as f:
            data = f.readlines()
        key = ''
        for line in data[1:-1]:
            key = key + line.replace('\n', '')
        logger.info('Get public key content: \n{}'.format(key))
        key_dict[sn] = [mac, key]
    return key_dict


def write_file(cityname, key_dict):
    outwb = load_workbook(xlsx)
    try:
        del outwb[cityname]
    except KeyError:
        pass
    sheet = outwb.create_sheet(0)
    sheet.title = cityname
    for index, sn in enumerate(key_dict):
        column = int(index)+1
        sheet['A{}'.format(column)].value = sn
        sheet['B{}'.format(column)].value = key_dict.get(sn)[0]
        sheet['C{}'.format(column)].value = key_dict.get(sn)[1]
    outwb.save(xlsx)
    logger.info('Save public key content success')

